# data_sims.py
from flask import (
    Blueprint, render_template, request, jsonify,
    send_file, session, redirect, url_for, flash
)
import pandas as pd
from functools import wraps
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from data_store import load_data  # cache global
import traceback
import uuid

data_sims_bp = Blueprint("data_sims", __name__)

TEMPLATE_FILE = "Template Format Pemeriksaan UPLOAD.xlsx"

# =====================
# CSRF TOKEN HELPER
# =====================
def get_csrf_token():
    """Generate CSRF token sesuai dengan app.py Anda"""
    if '_csrf_token' not in session:
        session['_csrf_token'] = str(uuid.uuid4())
    return session['_csrf_token']

def validate_csrf():
    """Validasi CSRF token untuk POST requests"""
    if request.method == "POST":
        token = session.pop('_csrf_token', None)
        received_token = request.form.get('csrf_token') or request.headers.get('X-CSRF-Token')
        
        # Allow API endpoints tanpa CSRF untuk DataTables
        if request.endpoint in ['data_sims.api', 'data_sims.download_post']:
            return True
            
        if not token or token != received_token:
            return False
    return True

# =====================
# LOGIN REQUIRED (sama seperti app.py)
# =====================
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            flash("Silakan login terlebih dahulu", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper

# =====================
# FILTER DATA - IMPROVED (PERBAIKAN STRING ACCESSOR)
# =====================
def apply_filter(df, search_value):
    if not search_value or df.empty:
        return df

    print(f"Applying filter for: '{search_value}'")

    try:
        # PERBAIKAN: Pastikan semua kolom adalah string
        df = df.astype(str)
        
        # Convert ke lowercase untuk pencarian
        df_search = df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
        search_lower = search_value.lower()

        # Split keywords
        keywords = search_lower.split()
        angka = [k for k in keywords if k.isdigit()]
        teks = [k for k in keywords if not k.isdigit()]

        # Buat mask untuk filter
        mask = pd.Series([True] * len(df), index=df.index)

        # Filter angka (exact match di cell manapun)
        if angka:
            angka_mask = pd.Series([False] * len(df), index=df.index)
            for a in angka:
                # Cari di semua kolom
                for col in df_search.columns:
                    angka_mask = angka_mask | (df_search[col] == a)
            mask = mask & angka_mask

        # Filter teks (partial match OR di kolom manapun)
        if teks:
            teks_mask = pd.Series([False] * len(df), index=df.index)
            for t in teks:
                for col in df_search.columns:
                    # PERBAIKAN: Gunakan .str.contains() dengan handling yang benar
                    col_data = df_search[col].astype(str)
                    teks_mask = teks_mask | col_data.str.contains(t, na=False)
            mask = mask & teks_mask

        result = df[mask].copy()
        print(f"Filter result: {len(result)} rows from {len(df)}")
        return result

    except Exception as e:
        print(f"Error in apply_filter: {e}")
        traceback.print_exc()
        return df

# =====================
# PAGE DATA SIMS
# =====================
@data_sims_bp.route("/data-sims", strict_slashes=False)
@login_required
def page():
    try:
        df = load_data()
        print(f"Data SIMS page - Loaded {len(df)} rows")

        # Hapus kolom 'no' untuk tampilan
        cols_to_remove = ['no', 'No', 'NO']
        columns = []
        for col in df.columns:
            if col not in cols_to_remove:
                columns.append(col)

        print(f"Display columns: {columns}")
        
        # Generate CSRF token untuk template
        csrf_token = get_csrf_token()
        
        return render_template("data_sims.html", columns=columns, csrf_token=csrf_token)
        
    except Exception as e:
        print(f"Error in data_sims page: {e}")
        flash(f"Error loading data: {str(e)}", "danger")
        return render_template("data_sims.html", columns=[], csrf_token=get_csrf_token())

# =====================
# API DATATABLES - FIXED (POST METHOD UNTUK 414 ERROR)
# =====================
@data_sims_bp.route("/data", methods=["POST"], strict_slashes=False)
@login_required
def api():
    """API for DataTables - menggunakan POST untuk menghindari 414 error"""
    try:
        # Validasi CSRF ringan untuk API
        if not request.is_json:
            token = request.form.get('csrf_token')
            if token and token != session.get('_csrf_token'):
                return jsonify({"error": "Invalid CSRF token"}), 403
        
        # Load data
        df = load_data()
        print(f"API called - Total rows: {len(df)}")

        if df.empty:
            print("DataFrame is empty")
            return jsonify({
                "draw": 1,
                "recordsTotal": 0,
                "recordsFiltered": 0,
                "data": []
            })

        # Hapus kolom no untuk display
        display_df = df.copy()
        cols_to_drop = ['no', 'No', 'NO']
        for col in cols_to_drop:
            if col in display_df.columns:
                display_df = display_df.drop(columns=[col])

        # Pastikan semua string
        display_df = display_df.astype(str)

        # Get Datatable parameters from POST data
        data = request.get_json() if request.is_json else request.form
        if not data:
            # Fallback to GET parameters jika tidak ada POST data
            data = request.args
        
        draw = int(data.get("draw", 1))
        start = int(data.get("start", 0))
        length = int(data.get("length", 10))
        
        # Handle search parameter (format berbeda-beda)
        search_value = ""
        if "search[value]" in data:
            search_value = data.get("search[value]", "").strip()
        elif "search_value" in data:
            search_value = data.get("search_value", "").strip()
        elif "search" in data:
            search_value = data.get("search", "").strip()
            
        print(f"API params - draw:{draw}, start:{start}, length:{length}, search:'{search_value[:50]}...'")

        # Apply filter
        filtered_df = apply_filter(display_df, search_value)

        # Paginate
        total_records = len(display_df)
        total_filtered = len(filtered_df)
        page_df = filtered_df.iloc[start:start + length] if total_filtered > 0 else pd.DataFrame()

        # Prepare data for Datatables
        data_list = []
        for _, row in page_df.iterrows():
            # Convert row ke list dalam urutan kolom
            row_data = [str(row[col]) if pd.notna(row[col]) else "" for col in display_df.columns]
            data_list.append(row_data)

        print(f"Returning {len(data_list)} rows, total:{total_records}, filtered:{total_filtered}")

        response = {
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": total_filtered,
            "data": data_list
        }

        return jsonify(response)

    except Exception as e:
        print(f"Error in data_sims API: {str(e)}")
        traceback.print_exc()
        return jsonify({
            "draw": 1,
            "recordsTotal": 0,
            "recordsFiltered": 0,
            "data": [],
            "error": str(e)
        }), 500

# =====================
# API ALTERNATIF UNTUK GET (UNTUK SEARCH PANJANG TERBATAS)
# =====================
@data_sims_bp.route("/data-get", methods=["GET"], strict_slashes=False)
@login_required
def api_get():
    """Alternatif API dengan GET untuk kompatibilitas"""
    try:
        # Load data
        df = load_data()
        
        if df.empty:
            return jsonify({
                "draw": 1,
                "recordsTotal": 0,
                "recordsFiltered": 0,
                "data": []
            })

        # Hapus kolom no untuk display
        display_df = df.copy()
        cols_to_drop = ['no', 'No', 'NO']
        for col in cols_to_drop:
            if col in display_df.columns:
                display_df = display_df.drop(columns=[col])

        # Pastikan semua string
        display_df = display_df.astype(str)

        # Get parameters
        draw = int(request.args.get("draw", 1))
        start = int(request.args.get("start", 0))
        length = int(request.args.get("length", 10))
        search_value = request.args.get("search[value]", "").strip()
        
        # Batasi panjang search untuk menghindari 414 error
        if len(search_value) > 1000:
            search_value = search_value[:1000]
            print(f"Search truncated to 1000 characters")

        # Apply filter
        filtered_df = apply_filter(display_df, search_value)

        # Paginate
        total_records = len(display_df)
        total_filtered = len(filtered_df)
        page_df = filtered_df.iloc[start:start + length] if total_filtered > 0 else pd.DataFrame()

        # Prepare data
        data_list = []
        for _, row in page_df.iterrows():
            row_data = [str(row[col]) if pd.notna(row[col]) else "" for col in display_df.columns]
            data_list.append(row_data)

        response = {
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": total_filtered,
            "data": data_list
        }

        return jsonify(response)

    except Exception as e:
        print(f"Error in data_sims API GET: {str(e)}")
        return jsonify({
            "draw": 1,
            "recordsTotal": 0,
            "recordsFiltered": 0,
            "data": [],
            "error": str(e)
        }), 500

# =====================
# HELPER
# =====================
def div_1000(val):
    try:
        return float(val) / 1000
    except:
        return ""

# =====================
# DOWNLOAD EXCEL VIA POST (FIX 414 ERROR)
# =====================
@data_sims_bp.route("/data-sims/download", methods=["POST"], strict_slashes=False)
@login_required
def download_post():
    try:
        # Validasi CSRF
        token = request.form.get('csrf_token')
        if token != session.get('_csrf_token'):
            flash("Invalid CSRF token", "danger")
            return redirect(url_for('data_sims.page'))
        
        search_value = request.form.get("search", "")
        print(f"Download POST - Search: '{search_value}'")

        df = load_data()

        # Hapus kolom no
        cols_to_drop = ['no', 'No', 'NO']
        for col in cols_to_drop:
            if col in df.columns:
                df = df.drop(columns=[col])

        # Apply filter jika ada search
        if search_value:
            df = apply_filter(df, search_value)
        
        print(f"Downloading {len(df)} rows")

        wb = load_workbook(TEMPLATE_FILE)
        ws = wb.active
        START_ROW = 7

        # Data validations
        dv_metode = DataValidation(
            type="list",
            formula1='"Inspeksi melalui Open Shelter,Pemeriksaan melalui Remote Site"',
            allow_blank=True
        )
        dv_sertifikat = DataValidation(
            type="list",
            formula1='"Ada,Tidak"',
            allow_blank=True
        )
        dv_status = DataValidation(
            type="list",
            formula1='"Sesuai ISR,Tidak Sesuai Parameter Teknis,Tidak Berizin,Tidak Aktif"',
            allow_blank=True
        )

        ws.add_data_validation(dv_metode)
        ws.add_data_validation(dv_sertifikat)
        ws.add_data_validation(dv_status)

        for i, (_, r) in enumerate(df.iterrows()):
            row = START_ROW + i

            ws.cell(row=row, column=1).value = i + 1
            dv_metode.add(ws.cell(row=row, column=3))

            # Fill data dengan handle None/NaN
            ws.cell(row=row, column=4).value = str(r.get("CLNT_ID", "")) if pd.notna(r.get("CLNT_ID")) else ""
            ws.cell(row=row, column=5).value = str(r.get("CLNT_NAME", "")) if pd.notna(r.get("CLNT_NAME")) else ""
            ws.cell(row=row, column=7).value = str(r.get("CURR_LIC_NUM", "")) if pd.notna(r.get("CURR_LIC_NUM")) else ""
            ws.cell(row=row, column=8).value = str(r.get("LINK_ID", "")) if pd.notna(r.get("LINK_ID")) else ""
            ws.cell(row=row, column=9).value = str(r.get("STN_NAME", "")) if pd.notna(r.get("STN_NAME")) else ""
            ws.cell(row=row, column=10).value = str(r.get("STASIUN_LAWAN", "")) if pd.notna(r.get("STASIUN_LAWAN")) else ""
            ws.cell(row=row, column=11).value = str(r.get("SID_LONG", "")) if pd.notna(r.get("SID_LONG")) else ""
            ws.cell(row=row, column=12).value = str(r.get("SID_LAT", "")) if pd.notna(r.get("SID_LAT")) else ""
            ws.cell(row=row, column=13).value = str(r.get("FREQ", "")) if pd.notna(r.get("FREQ")) else ""
            ws.cell(row=row, column=14).value = str(r.get("FREQ_PAIR", "")) if pd.notna(r.get("FREQ_PAIR")) else ""
            bwidth = str(r.get("BWIDTH", "")) if pd.notna(r.get("BWIDTH")) else ""
            ws.cell(row=row, column=15).value = div_1000(bwidth)
            ws.cell(row=row, column=16).value = str(r.get("EQ_MDL", "")) if pd.notna(r.get("EQ_MDL")) else ""
            ws.cell(row=row, column=17).value = str(r.get("STN_NAME", "")) if pd.notna(r.get("STN_NAME")) else ""
            ws.cell(row=row, column=18).value = str(r.get("STASIUN_LAWAN", "")) if pd.notna(r.get("STASIUN_LAWAN")) else ""
            ws.cell(row=row, column=19).value = str(r.get("LONG", "")) if pd.notna(r.get("LONG")) else ""
            ws.cell(row=row, column=20).value = str(r.get("LAT", "")) if pd.notna(r.get("LAT")) else ""
            ws.cell(row=row, column=21).value = str(r.get("FREQ", "")) if pd.notna(r.get("FREQ")) else ""
            ws.cell(row=row, column=22).value = str(r.get("FREQ_PAIR", "")) if pd.notna(r.get("FREQ_PAIR")) else ""
            ws.cell(row=row, column=23).value = div_1000(bwidth)
            ws.cell(row=row, column=24).value = str(r.get("EQ_MDL", "")) if pd.notna(r.get("EQ_MDL")) else ""

            dv_sertifikat.add(ws.cell(row=row, column=25))
            dv_status.add(ws.cell(row=row, column=26))
            ws.cell(row=row, column=27).value = str(r.get("MULAI BEROPERASI", "")) if pd.notna(r.get("MULAI BEROPERASI")) else ""
            ws.cell(row=row, column=28).value = str(r.get("KETERANGAN", "")) if pd.notna(r.get("KETERANGAN")) else ""
            ws.cell(row=row, column=29).value = str(r.get("CITY", "")) if pd.notna(r.get("CITY")) else ""

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
        filename = f"laporan_data_sims_{timestamp}.xlsx"

        return send_file(
            out,
            download_name=filename,
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Error downloading file via POST: {e}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# =====================
# DOWNLOAD SIMPLE (TANPA FILTER)
# =====================
@data_sims_bp.route("/data-sims/download-all", strict_slashes=False)
@login_required
def download_all():
    """Download semua data tanpa filter"""
    try:
        df = load_data()

        # Hapus kolom no
        cols_to_drop = ['no', 'No', 'NO']
        for col in cols_to_drop:
            if col in df.columns:
                df = df.drop(columns=[col])
        
        print(f"Downloading ALL {len(df)} rows")

        wb = load_workbook(TEMPLATE_FILE)
        ws = wb.active
        START_ROW = 7

        # Data validations
        dv_metode = DataValidation(
            type="list",
            formula1='"Inspeksi melalui Open Shelter,Pemeriksaan melalui Remote Site"',
            allow_blank=True
        )
        dv_sertifikat = DataValidation(
            type="list",
            formula1='"Ada,Tidak"',
            allow_blank=True
        )
        dv_status = DataValidation(
            type="list",
            formula1='"Sesuai ISR,Tidak Sesuai Parameter Teknis,Tidak Berizin,Tidak Aktif"',
            allow_blank=True
        )

        ws.add_data_validation(dv_metode)
        ws.add_data_validation(dv_sertifikat)
        ws.add_data_validation(dv_status)

        for i, (_, r) in enumerate(df.iterrows()):
            row = START_ROW + i

            ws.cell(row=row, column=1).value = i + 1
            dv_metode.add(ws.cell(row=row, column=3))

            # Fill data
            ws.cell(row=row, column=4).value = str(r.get("CLNT_ID", "")) if pd.notna(r.get("CLNT_ID")) else ""
            ws.cell(row=row, column=5).value = str(r.get("CLNT_NAME", "")) if pd.notna(r.get("CLNT_NAME")) else ""
            ws.cell(row=row, column=7).value = str(r.get("CURR_LIC_NUM", "")) if pd.notna(r.get("CURR_LIC_NUM")) else ""
            ws.cell(row=row, column=8).value = str(r.get("LINK_ID", "")) if pd.notna(r.get("LINK_ID")) else ""
            ws.cell(row=row, column=9).value = str(r.get("STN_NAME", "")) if pd.notna(r.get("STN_NAME")) else ""
            ws.cell(row=row, column=10).value = str(r.get("STASIUN_LAWAN", "")) if pd.notna(r.get("STASIUN_LAWAN")) else ""
            ws.cell(row=row, column=11).value = str(r.get("SID_LONG", "")) if pd.notna(r.get("SID_LONG")) else ""
            ws.cell(row=row, column=12).value = str(r.get("SID_LAT", "")) if pd.notna(r.get("SID_LAT")) else ""
            ws.cell(row=row, column=13).value = str(r.get("FREQ", "")) if pd.notna(r.get("FREQ")) else ""
            ws.cell(row=row, column=14).value = str(r.get("FREQ_PAIR", "")) if pd.notna(r.get("FREQ_PAIR")) else ""
            bwidth = str(r.get("BWIDTH", "")) if pd.notna(r.get("BWIDTH")) else ""
            ws.cell(row=row, column=15).value = div_1000(bwidth)
            ws.cell(row=row, column=16).value = str(r.get("EQ_MDL", "")) if pd.notna(r.get("EQ_MDL")) else ""
            ws.cell(row=row, column=17).value = str(r.get("STN_NAME", "")) if pd.notna(r.get("STN_NAME")) else ""
            ws.cell(row=row, column=18).value = str(r.get("STASIUN_LAWAN", "")) if pd.notna(r.get("STASIUN_LAWAN")) else ""
            ws.cell(row=row, column=19).value = str(r.get("LONG", "")) if pd.notna(r.get("LONG")) else ""
            ws.cell(row=row, column=20).value = str(r.get("LAT", "")) if pd.notna(r.get("LAT")) else ""
            ws.cell(row=row, column=21).value = str(r.get("FREQ", "")) if pd.notna(r.get("FREQ")) else ""
            ws.cell(row=row, column=22).value = str(r.get("FREQ_PAIR", "")) if pd.notna(r.get("FREQ_PAIR")) else ""
            ws.cell(row=row, column=23).value = div_1000(bwidth)
            ws.cell(row=row, column=24).value = str(r.get("EQ_MDL", "")) if pd.notna(r.get("EQ_MDL")) else ""

            dv_sertifikat.add(ws.cell(row=row, column=25))
            dv_status.add(ws.cell(row=row, column=26))
            ws.cell(row=row, column=27).value = str(r.get("MULAI BEROPERASI", "")) if pd.notna(r.get("MULAI BEROPERASI")) else ""
            ws.cell(row=row, column=28).value = str(r.get("KETERANGAN", "")) if pd.notna(r.get("KETERANGAN")) else ""
            ws.cell(row=row, column=29).value = str(r.get("CITY", "")) if pd.notna(r.get("CITY")) else ""

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
        filename = f"laporan_data_sims_all_{timestamp}.xlsx"

        return send_file(
            out,
            download_name=filename,
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Error downloading all data: {e}")
        traceback.print_exc()
        flash(f"Error downloading file: {str(e)}", "danger")
        return redirect(url_for('data_sims.page'))