from flask import Blueprint, render_template, request, jsonify, send_file, redirect, url_for
import pandas as pd
import os, re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from data_store import load_data  # cache global
import numpy as np
import traceback

pemeriksaan_bp = Blueprint("pemeriksaan", __name__)

TEMPLATE_FILE = "Template Format Pemeriksaan UPLOAD.xlsx"
SAVED_FILE = "data_pemeriksaan_tersimpan.xlsx"

DISPLAY_COLUMNS = [
    "CLNT_ID", "CLNT_NAME", "CURR_LIC_NUM",
    "LINK_ID", "STN_NAME", "STASIUN_LAWAN",
    "SID_LONG", "SID_LAT", "FREQ",
    "FREQ_PAIR", "BWIDTH", "EQ_MDL", "CITY"
]

FIELD_MAP = {
    "client_id": "CLNT_ID",
    "client_name": "CLNT_NAME",
    "link_id": "LINK_ID",
    "stn_name": "STN_NAME",
    "stasiun_lawan": "STASIUN_LAWAN",
    "freq": "FREQ",
    "city": "CITY"
}

# =======================
def apply_filter(args):
    df = load_data()
    for field, col in FIELD_MAP.items():
        val = args.get(field, "").strip()
        if val:
            parts = [v.strip() for v in val.split(";")]
            pattern = "|".join(map(re.escape, parts))
            df = df[df[col].str.contains(pattern, case=False, na=False)]
    return df

# =======================
def div_1000(val):
    try:
        return float(val) / 1000
    except:
        return ""

# =======================
def generate_excel(data, filename):
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active
    START_ROW = 7

    dv_metode = DataValidation(type="list",
        formula1='"Inspeksi melalui Open Shelter,Pemeriksaan melalui Remote Site"')
    dv_sertifikat = DataValidation(type="list", formula1='"Ada,Tidak"')
    dv_status = DataValidation(type="list",
        formula1='"Sesuai ISR,Tidak Sesuai Parameter Teknis,Tidak Berizin,Tidak Aktif"')

    ws.add_data_validation(dv_metode)
    ws.add_data_validation(dv_sertifikat)
    ws.add_data_validation(dv_status)

    for i, r in enumerate(data):
        row = START_ROW + i
        ws.cell(row=row, column=1).value = i + 1
        dv_metode.add(ws.cell(row=row, column=3))

        ws.cell(row=row, column=4).value = r.get("CLNT_ID", "") or ""
        ws.cell(row=row, column=5).value = r.get("CLNT_NAME", "") or ""
        ws.cell(row=row, column=7).value = r.get("CURR_LIC_NUM", "") or ""
        ws.cell(row=row, column=8).value = r.get("LINK_ID", "") or ""
        ws.cell(row=row, column=9).value = r.get("STN_NAME", "") or ""
        ws.cell(row=row, column=10).value = r.get("STASIUN_LAWAN", "") or ""
        ws.cell(row=row, column=11).value = r.get("SID_LONG", "") or ""
        ws.cell(row=row, column=12).value = r.get("SID_LAT", "") or ""
        ws.cell(row=row, column=13).value = r.get("FREQ", "") or ""
        ws.cell(row=row, column=14).value = r.get("FREQ_PAIR", "") or ""
        bwidth = r.get("BWIDTH", "") or ""
        ws.cell(row=row, column=15).value = div_1000(bwidth)
        ws.cell(row=row, column=16).value = r.get("EQ_MDL", "") or ""
        ws.cell(row=row, column=17).value = r.get("STN_NAME", "") or ""
        ws.cell(row=row, column=18).value = r.get("STASIUN_LAWAN", "") or ""
        ws.cell(row=row, column=19).value = r.get("LONG", "") or ""
        ws.cell(row=row, column=20).value = r.get("LAT", "") or ""
        ws.cell(row=row, column=21).value = r.get("FREQ", "") or ""
        ws.cell(row=row, column=22).value = r.get("FREQ_PAIR", "") or ""
        ws.cell(row=row, column=23).value = div_1000(bwidth)
        ws.cell(row=row, column=24).value = r.get("EQ_MDL", "") or ""

        dv_sertifikat.add(ws.cell(row=row, column=25))
        dv_status.add(ws.cell(row=row, column=26))
        ws.cell(row=row, column=27).value = r.get("MULAI BEROPERASI", "") or ""
        ws.cell(row=row, column=28).value = r.get("KETERANGAN", "") or ""
        ws.cell(row=row, column=29).value = r.get("CITY", "") or ""

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, download_name=filename, as_attachment=True)

# =======================
def clean_value(value):
    """Membersihkan nilai: NaN/None menjadi string kosong"""
    if pd.isna(value):
        return ""
    if value is None:
        return ""
    if value == "nan" or value == "None" or value == "null":
        return ""
    # Jika sudah string, kembalikan as-is
    if isinstance(value, str):
        return value.strip()
    # Jika numeric, konversi ke string
    try:
        return str(value).strip()
    except:
        return ""

def is_row_duplicate(row1, row2):
    """Perbandingan semua kolom termasuk yang kosong"""
    # Daftar semua kolom yang relevan untuk perbandingan
    columns_to_compare = DISPLAY_COLUMNS

    for col in columns_to_compare:
        val1 = clean_value(row1.get(col, ''))
        val2 = clean_value(row2.get(col, ''))

        # Bandingkan nilai yang sudah dibersihkan
        if val1 != val2:
            return False

    return True

# =======================
def prepare_dataframe(df):
    """Prepare DataFrame: clean NaN values and ensure proper columns"""
    # Pastikan semua kolom ada
    for col in DISPLAY_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    # Replace NaN dengan string kosong
    df = df.fillna("")

    # Convert semua kolom ke string untuk konsistensi
    for col in DISPLAY_COLUMNS:
        df[col] = df[col].astype(str)
        # Hapus whitespace berlebih dan clean nilai
        df[col] = df[col].apply(lambda x: clean_value(x))

    return df

# =======================
# ENDPOINT UNTUK EDIT DAN DELETE SATU DATA
# =======================
@pemeriksaan_bp.route("/pemeriksaan/delete_single", methods=["POST"], strict_slashes=False)
def delete_single():
    try:
        data = request.json
        row_index = data.get("index")

        print(f"DELETE SINGLE - Index: {row_index}")

        if not os.path.exists(SAVED_FILE):
            return jsonify({
                "success": False,
                "message": "File data tidak ditemukan"
            }), 404

        # Baca data
        df = pd.read_excel(SAVED_FILE, dtype=str)
        df = prepare_dataframe(df)

        # Validasi index
        if row_index is None or row_index < 0 or row_index >= len(df):
            return jsonify({
                "success": False,
                "message": f"Index tidak valid: {row_index}. Data hanya {len(df)} baris"
            }), 400

        print(f"Data sebelum hapus: {len(df)} baris")
        print(f"Menghapus baris ke-{row_index}")

        # Simpan data yang akan dihapus untuk debug
        deleted_row = df.iloc[row_index].to_dict()
        print(f"Data yang dihapus: {deleted_row}")

        # Hapus baris berdasarkan index
        df = df.drop(index=row_index).reset_index(drop=True)

        print(f"Data setelah hapus: {len(df)} baris")

        # Simpan kembali
        df = prepare_dataframe(df)

        # Simpan ke file dengan engine yang jelas
        try:
            df.to_excel(SAVED_FILE, index=False, engine='openpyxl')
            print("File berhasil disimpan ulang")
        except Exception as save_error:
            print(f"Error saat menyimpan: {save_error}")
            # Coba dengan engine lain
            df.to_excel(SAVED_FILE, index=False)

        # Verifikasi file tersimpan
        if os.path.exists(SAVED_FILE):
            df_check = pd.read_excel(SAVED_FILE, dtype=str)
            print(f"Verifikasi: file sekarang berisi {len(df_check)} baris")

        return jsonify({
            "success": True,
            "message": "Data berhasil dihapus",
            "remaining_count": len(df),
            "deleted_index": row_index
        }), 200

    except Exception as e:
        print(f"Error deleting single row: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Terjadi kesalahan: {str(e)}"
        }), 500

# =======================
@pemeriksaan_bp.route("/pemeriksaan/update_single", methods=["POST"], strict_slashes=False)
def update_single():
    try:
        data = request.json
        row_index = data.get("index")
        updated_data = data.get("data", {})

        print(f"Updating row index: {row_index}")
        print(f"Updated data: {updated_data}")

        if not os.path.exists(SAVED_FILE):
            return jsonify({
                "success": False,
                "message": "File data tidak ditemukan"
            }), 404

        # Baca data
        df = pd.read_excel(SAVED_FILE, dtype=str)
        df = prepare_dataframe(df)

        # Validasi index
        if row_index is None or row_index < 0 or row_index >= len(df):
            return jsonify({
                "success": False,
                "message": f"Index tidak valid: {row_index}. Data hanya {len(df)} baris"
            }), 400

        # Update semua kolom (termasuk yang kosong)
        for col in DISPLAY_COLUMNS:
            if col in updated_data:
                df.at[row_index, col] = clean_value(updated_data[col])
            else:
                # Jika kolom tidak ada di updated_data, set ke string kosong
                df.at[row_index, col] = ""

        # Simpan kembali
        df = prepare_dataframe(df)
        df.to_excel(SAVED_FILE, index=False)

        # Ambil data yang sudah diupdate untuk response
        updated_row = df.iloc[row_index].to_dict()

        return jsonify({
            "success": True,
            "message": "Data berhasil diperbarui",
            "data": updated_row
        }), 200

    except Exception as e:
        print(f"Error updating single row: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": f"Terjadi kesalahan: {str(e)}"
        }), 500

# =======================
# ROUTE UTAMA
# =======================
@pemeriksaan_bp.route("/pemeriksaan/", strict_slashes=False)
def index():
    return render_template("pemeriksaan.html")

@pemeriksaan_bp.route("/pemeriksaan/api", strict_slashes=False)
def api():
    df = apply_filter(request.args)
    draw = int(request.args.get("draw", 1))
    start = int(request.args.get("start", 0))
    length = int(request.args.get("length", 10))

    page = df.iloc[start:start+length]
    data = [["", *r[DISPLAY_COLUMNS].tolist()] for _, r in page.iterrows()]

    return jsonify({
        "draw": draw,
        "recordsTotal": len(load_data()),
        "recordsFiltered": len(df),
        "data": data
    })

@pemeriksaan_bp.route("/pemeriksaan/save", methods=["POST"], strict_slashes=False)
def save_selected():
    try:
        rows = request.json.get("rows", [])

        if not rows:
            return jsonify({
                "message": "Tidak ada data yang dipilih",
                "status": "warning"
            }), 200

        # Clean rows data: replace None dengan string kosong
        cleaned_rows = []
        for row in rows:
            cleaned_row = {}
            for col in DISPLAY_COLUMNS:
                value = row.get(col, '')
                cleaned_row[col] = clean_value(value)
            cleaned_rows.append(cleaned_row)

        # Convert to DataFrame dengan data yang sudah dibersihkan
        df_new = pd.DataFrame(cleaned_rows)

        # Jika tidak ada file tersimpan, buat baru
        if not os.path.exists(SAVED_FILE):
            df_new = prepare_dataframe(df_new)
            df_new.to_excel(SAVED_FILE, index=False)
            return jsonify({
                "message": f"{len(rows)} data disimpan",
                "status": "success",
                "new_count": len(rows),
                "duplicate_count": 0,
                "redirect": url_for('pemeriksaan.saved_page', saved=1, count=len(rows))
            }), 200

        # Load data yang sudah tersimpan
        try:
            df_old = pd.read_excel(SAVED_FILE, dtype=str)  # Baca semua sebagai string
            df_old = prepare_dataframe(df_old)  # Clean NaN values
        except Exception as e:
            # Jika file corrupt, buat baru
            df_new = prepare_dataframe(df_new)
            df_new.to_excel(SAVED_FILE, index=False)
            return jsonify({
                "message": f"{len(rows)} data disimpan (file baru dibuat)",
                "status": "success",
                "new_count": len(rows),
                "duplicate_count": 0,
                "redirect": url_for('pemeriksaan.saved_page', saved=1, count=len(rows))
            }), 200

        # Cek duplikat
        new_data_rows = []
        duplicate_rows_count = 0

        for _, new_row in df_new.iterrows():
            is_duplicate = False
            new_row_dict = new_row.to_dict()

            # Cek duplikat dengan data yang sudah ada
            for _, old_row in df_old.iterrows():
                if is_row_duplicate(new_row_dict, old_row.to_dict()):
                    is_duplicate = True
                    duplicate_rows_count += 1
                    break

            # Jika bukan duplikat, tambahkan ke list baru
            if not is_duplicate:
                new_data_rows.append(new_row_dict)

        # Gabungkan data lama dengan data baru (non-duplikat)
        if new_data_rows:
            df_new_filtered = pd.DataFrame(new_data_rows)
            df_combined = pd.concat([df_old, df_new_filtered], ignore_index=True)

            # Simpan dengan format yang benar (tanpa NaN)
            df_combined = prepare_dataframe(df_combined)
            df_combined.to_excel(SAVED_FILE, index=False)

        # Hitung berapa data baru yang tersimpan
        new_rows_count = len(new_data_rows)

        # Prepare response berdasarkan skenario
        if duplicate_rows_count > 0 and new_rows_count == 0:
            return jsonify({
                "message": f"{duplicate_rows_count} Data Sudah Ada Tersimpan",
                "status": "duplicate_all",
                "new_count": 0,
                "duplicate_count": duplicate_rows_count
            }), 200

        elif duplicate_rows_count > 0 and new_rows_count > 0:
            return jsonify({
                "message": f"{new_rows_count} data disimpan, {duplicate_rows_count} data sudah ada",
                "status": "partial",
                "new_count": new_rows_count,
                "duplicate_count": duplicate_rows_count,
                "redirect": url_for('pemeriksaan.saved_page', saved=1, count=new_rows_count)
            }), 200

        else:
            return jsonify({
                "message": f"{new_rows_count} data disimpan",
                "status": "success",
                "new_count": new_rows_count,
                "duplicate_count": 0,
                "redirect": url_for('pemeriksaan.saved_page', saved=1, count=new_rows_count)
            }), 200

    except Exception as e:
        # Log error untuk debugging
        print(f"Error in save_selected: {str(e)}")
        print(traceback.format_exc())

        return jsonify({
            "message": f"Terjadi kesalahan server: {str(e)}",
            "status": "error"
        }), 500

# =======================
# HALAMAN DATA TERSIMPAN
# =======================
@pemeriksaan_bp.route("/pemeriksaan/saved", strict_slashes=False)
def saved_page():
    data = []
    if os.path.exists(SAVED_FILE):
        try:
            df = pd.read_excel(SAVED_FILE, dtype=str)
            df = prepare_dataframe(df)
            data = df[DISPLAY_COLUMNS].values.tolist()
        except Exception as e:
            print(f"Error loading saved data: {e}")

    # Kirim DISPLAY_COLUMNS ke template
    return render_template(
        "pemeriksaan_saved.html",
        data=data,
        DISPLAY_COLUMNS=DISPLAY_COLUMNS
    )

# =======================
# DOWNLOAD & CLEAR
# =======================
@pemeriksaan_bp.route("/pemeriksaan/download_saved", strict_slashes=False)
def download_saved():
    if not os.path.exists(SAVED_FILE):
        return redirect(url_for("pemeriksaan.saved_page"))

    try:
        df = pd.read_excel(SAVED_FILE, dtype=str)
        df = prepare_dataframe(df)
        records = df.to_dict('records')
        return generate_excel(records, "data_pemeriksaan_tersimpan.xlsx")
    except Exception as e:
        print(f"Error downloading saved data: {e}")
        return redirect(url_for("pemeriksaan.saved_page"))

# PERBAIKAN: Tambahkan parameter untuk alert
@pemeriksaan_bp.route("/pemeriksaan/clear", strict_slashes=False)
def clear_saved():
    deleted_count = 0

    if os.path.exists(SAVED_FILE):
        try:
            # Hitung berapa data yang dihapus
            df = pd.read_excel(SAVED_FILE)
            deleted_count = len(df)
        except:
            pass

        # Hapus file
        os.remove(SAVED_FILE)

        # Redirect dengan parameter untuk alert
        return redirect(url_for('pemeriksaan.saved_page', deleted=1, count=deleted_count))

    return redirect(url_for("pemeriksaan.saved_page"))

@pemeriksaan_bp.route("/pemeriksaan/download-filtered", strict_slashes=False)
def download_filtered():
    df = apply_filter(request.args)
    df = prepare_dataframe(df)
    records = df.to_dict('records')
    return generate_excel(records, "hasil_filter.xlsx")